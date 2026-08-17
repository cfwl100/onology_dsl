#!/usr/bin/env python3
"""OAG ontology embedding structure evaluation tool.

The tool reads ``OAG_本体索引结构.xlsx``, builds several vector-document
layouts, embeds them with BGE-M3 (1024 dimensions), writes the vectors to
GaussVector, runs segmented-query retrieval, and exports comparable metrics.

An in-memory exact cosine backend and a deterministic hash embedder are kept
for pipeline tests. They are not substitutes for the production BGE-M3 +
GaussVector experiment.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import statistics
import time
import unicodedata
import urllib.request
import uuid
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Protocol, Sequence

import numpy as np


DIMENSION = 1024
DEFAULT_VARIANTS = [
    "all_fields_single",
    "core_short_single",
    "term_description_multi",
    "context_shadow_multi",
    "atomic_field_rows",
]
DEFAULT_QUERY_STRATEGIES = ["whole", "token", "semantic_phrase", "hybrid"]

SEED_TERM_FIELDS = [
    "name",
    "display_zh",
    "display_en",
    "display_lang_1",
    "display_lang_2",
    "synonyms",
]
SEED_DESCRIPTION_FIELDS = [
    "description_zh",
    "description_en",
    "description_lang_1",
    "description_lang_2",
]
METADATA_TERM_FIELDS = [
    "value",
    "name",
    "display_zh",
    "display_en",
    "display_lang_1",
    "display_lang_2",
    "synonyms",
]
METADATA_DESCRIPTION_FIELDS = SEED_DESCRIPTION_FIELDS


@dataclass(frozen=True)
class EntityRecord:
    entity_type: str
    entity_key: str
    group_key: str
    fields: dict[str, str]


@dataclass(frozen=True)
class VectorDocument:
    variant: str
    entity_type: str
    entity_key: str
    group_key: str
    vector_role: str
    segment_no: int
    input_text: str


@dataclass(frozen=True)
class QueryCase:
    query_id: str
    question: str
    expected_entity_keys: tuple[str, ...]
    semantic_units: tuple[str, ...] = ()
    category: str = "unspecified"
    language: str = "und"
    origin: str = "file"


@dataclass(frozen=True)
class SearchHit:
    entity_key: str
    group_key: str
    entity_type: str
    vector_role: str
    similarity: float


@dataclass
class DataQuality:
    sheet_rows: dict[str, int] = field(default_factory=dict)
    loaded_entities: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    skipped_empty_values: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    duplicate_entity_keys: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    missing_fields: dict[str, dict[str, int]] = field(default_factory=dict)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def normalize_term(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    return re.sub(r"\s+", " ", value).strip().casefold()


def dedupe_parts(parts: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for part in parts:
        part = _text(part)
        if not part:
            continue
        key = normalize_term(part)
        if key and key not in seen:
            result.append(part)
            seen.add(key)
    return result


def split_synonyms(value: str) -> list[str]:
    return dedupe_parts(re.split(r"[\r\n]+", _text(value)))


def compose(fields: dict[str, str], names: Sequence[str], split_synonym: bool = False) -> str:
    parts: list[str] = []
    for name in names:
        value = fields.get(name, "")
        if name == "synonyms" and split_synonym:
            parts.extend(split_synonyms(value))
        elif value:
            parts.append(value)
    return "\n".join(dedupe_parts(parts))


def humanize_identifier(value: str) -> str:
    tail = re.split(r"[-:/]", value or "")[-1]
    tail = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", tail)
    tail = tail.replace("_", " ")
    return re.sub(r"\s+", " ", tail).strip()


def seed_key(identifier: str) -> str:
    return f"seed|{identifier}"


def metadata_key(property_id: str, object_type_id: str, value: str) -> str:
    return f"metadata|{property_id}|{object_type_id}|{value}"


def instance_key(property_id: str, object_type_id: str, value: str) -> str:
    return f"instance|{property_id}|{object_type_id}|{value}"


def load_workbook(path: Path) -> tuple[list[EntityRecord], DataQuality]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency guidance
        raise RuntimeError("Reading xlsx requires openpyxl. Install the evaluation requirements.") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    quality = DataQuality()
    entities: list[EntityRecord] = []
    seen: dict[str, set[str]] = defaultdict(set)

    expected = {
        "种子节点": ["type", "id", "parent_id", *SEED_TERM_FIELDS, *SEED_DESCRIPTION_FIELDS],
        "元数据元素": ["type", "propertyId", "objectTypeId", *METADATA_TERM_FIELDS, *METADATA_DESCRIPTION_FIELDS],
        "实例元素": ["type", "propertyid", "objectTypeId", "value"],
    }
    for sheet_name, required_columns in expected.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows)]
        absent = [column for column in required_columns if column not in headers]
        if absent:
            raise ValueError(f"Sheet {sheet_name} is missing columns: {absent}")
        quality.sheet_rows[sheet_name] = max(0, sheet.max_row - 1)
        missing = defaultdict(int)

        for row_number, values in enumerate(rows, start=2):
            row = {header: _text(value) for header, value in zip(headers, values)}
            for column in required_columns:
                if not row.get(column):
                    missing[column] += 1

            if sheet_name == "种子节点":
                identifier = row["id"]
                if not identifier:
                    quality.skipped_empty_values["seed.id"] += 1
                    continue
                entity = EntityRecord(
                    entity_type="seed",
                    entity_key=seed_key(identifier),
                    group_key=identifier,
                    fields=row,
                )
            elif sheet_name == "元数据元素":
                property_id, object_type_id, value = (
                    row["propertyId"],
                    row["objectTypeId"],
                    row["value"],
                )
                if not property_id or not value:
                    quality.skipped_empty_values["metadata.key"] += 1
                    continue
                entity = EntityRecord(
                    entity_type="metadata",
                    entity_key=metadata_key(property_id, object_type_id, value),
                    group_key=property_id,
                    fields=row,
                )
            else:
                property_id, object_type_id, value = (
                    row["propertyid"],
                    row["objectTypeId"],
                    row["value"],
                )
                if not value:
                    quality.skipped_empty_values["instance.value"] += 1
                    continue
                entity = EntityRecord(
                    entity_type="instance",
                    entity_key=instance_key(property_id, object_type_id, value),
                    group_key=property_id,
                    fields=row,
                )

            if entity.entity_key in seen[entity.entity_type]:
                quality.duplicate_entity_keys[entity.entity_type] += 1
                continue
            seen[entity.entity_type].add(entity.entity_key)
            quality.loaded_entities[entity.entity_type] += 1
            entities.append(entity)
        quality.missing_fields[sheet_name] = dict(missing)

    quality.loaded_entities = dict(quality.loaded_entities)
    quality.skipped_empty_values = dict(quality.skipped_empty_values)
    quality.duplicate_entity_keys = dict(quality.duplicate_entity_keys)
    return entities, quality


def build_documents(entities: Sequence[EntityRecord], variant: str) -> list[VectorDocument]:
    if variant not in DEFAULT_VARIANTS:
        raise ValueError(f"Unknown vectorization variant: {variant}")
    seed_by_id = {
        entity.fields.get("id", ""): entity
        for entity in entities
        if entity.entity_type == "seed"
    }
    documents: list[VectorDocument] = []

    def add(entity: EntityRecord, role: str, text: str, segment_no: int = 0) -> None:
        text = "\n".join(dedupe_parts(text.splitlines()))
        if text:
            documents.append(
                VectorDocument(
                    variant=variant,
                    entity_type=entity.entity_type,
                    entity_key=entity.entity_key,
                    group_key=entity.group_key,
                    vector_role=role,
                    segment_no=segment_no,
                    input_text=text,
                )
            )

    for entity in entities:
        fields = entity.fields
        if entity.entity_type == "seed":
            term_fields, description_fields = SEED_TERM_FIELDS, SEED_DESCRIPTION_FIELDS
        elif entity.entity_type == "metadata":
            term_fields, description_fields = METADATA_TERM_FIELDS, METADATA_DESCRIPTION_FIELDS
        else:
            term_fields, description_fields = ["value"], []

        term_text = compose(fields, term_fields, split_synonym=True)
        description_text = compose(fields, description_fields)
        all_text = compose(fields, [*term_fields[:-1], *description_fields, term_fields[-1]], True)

        if variant == "all_fields_single":
            add(entity, "all", all_text)
        elif variant == "core_short_single":
            add(entity, "term", term_text)
        elif variant == "term_description_multi":
            add(entity, "term", term_text)
            add(entity, "description", description_text)
        elif variant == "context_shadow_multi":
            add(entity, "term", term_text)
            add(entity, "description", description_text)
            context_parts: list[str] = []
            if entity.entity_type == "seed" and fields.get("parent_id"):
                parent = seed_by_id.get(fields["parent_id"])
                if parent:
                    context_parts.append(compose(parent.fields, SEED_TERM_FIELDS, True))
                context_parts.append(term_text)
            elif entity.entity_type in {"metadata", "instance"}:
                identifier = fields.get("propertyId") or fields.get("propertyid") or ""
                context_parts.extend([humanize_identifier(identifier), term_text])
            add(entity, "context", "\n".join(context_parts))
        elif variant == "atomic_field_rows":
            segment_no = 0
            for field_name in [*term_fields, *description_fields]:
                value = fields.get(field_name, "")
                values = split_synonyms(value) if field_name == "synonyms" else [value]
                for atomic_value in values:
                    add(entity, field_name, atomic_value, segment_no)
                    segment_no += 1
    return documents


class EmbeddingProvider(Protocol):
    dimension: int

    def embed(self, texts: Sequence[str]) -> np.ndarray: ...


class HashEmbeddingProvider:
    """Deterministic sparse hash vectors for unit tests only."""

    def __init__(self, dimension: int = DIMENSION) -> None:
        self.dimension = dimension

    def embed(self, texts: Sequence[str]) -> np.ndarray:
        output = np.zeros((len(texts), self.dimension), dtype=np.float32)
        for row, text in enumerate(texts):
            tokens = re.findall(r"[\w]+|[\u3400-\u9fff]", normalize_term(text))
            for token in tokens or [text]:
                digest = hashlib.blake2b(token.encode("utf-8"), digest_size=16).digest()
                index = int.from_bytes(digest[:8], "big") % self.dimension
                sign = 1.0 if digest[8] & 1 else -1.0
                output[row, index] += sign
        return normalize_vectors(output)


class SentenceTransformerProvider:
    def __init__(self, config: dict[str, Any]) -> None:
        try:
            from sentence_transformers import SentenceTransformer
        except ImportError as exc:  # pragma: no cover - dependency guidance
            raise RuntimeError("BGE-M3 requires sentence-transformers.") from exc
        self.dimension = int(config.get("dimension", DIMENSION))
        model_name = config.get("model_name_or_path", "BAAI/bge-m3")
        self.batch_size = int(config.get("batch_size", 32))
        self.model = SentenceTransformer(
            model_name,
            device=config.get("device"),
            trust_remote_code=bool(config.get("trust_remote_code", False)),
        )

    def embed(self, texts: Sequence[str]) -> np.ndarray:
        vectors = self.model.encode(
            list(texts),
            batch_size=self.batch_size,
            normalize_embeddings=True,
            show_progress_bar=False,
            convert_to_numpy=True,
        )
        vectors = np.asarray(vectors, dtype=np.float32)
        validate_dimension(vectors, self.dimension)
        return vectors


class HttpEmbeddingProvider:
    """OpenAI-compatible embeddings endpoint serving BGE-M3."""

    def __init__(self, config: dict[str, Any]) -> None:
        self.dimension = int(config.get("dimension", DIMENSION))
        self.endpoint = str(config["endpoint"])
        self.model = str(config.get("model_name_or_path", "BAAI/bge-m3"))
        self.batch_size = int(config.get("batch_size", 32))
        self.timeout = float(config.get("timeout_seconds", 120))
        token_env = str(config.get("token_env", "EMBEDDING_API_TOKEN"))
        self.token = os.getenv(token_env, "")

    def embed(self, texts: Sequence[str]) -> np.ndarray:
        all_vectors: list[list[float]] = []
        for start in range(0, len(texts), self.batch_size):
            payload = json.dumps(
                {"model": self.model, "input": list(texts[start : start + self.batch_size])}
            ).encode("utf-8")
            headers = {"Content-Type": "application/json"}
            if self.token:
                headers["Authorization"] = f"Bearer {self.token}"
            request = urllib.request.Request(self.endpoint, data=payload, headers=headers, method="POST")
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                result = json.loads(response.read().decode("utf-8"))
            all_vectors.extend(item["embedding"] for item in sorted(result["data"], key=lambda x: x["index"]))
        vectors = normalize_vectors(np.asarray(all_vectors, dtype=np.float32))
        validate_dimension(vectors, self.dimension)
        return vectors


class CachedEmbeddingProvider:
    """Avoid re-embedding identical index text and repeated query segments."""

    def __init__(self, delegate: EmbeddingProvider) -> None:
        self.delegate = delegate
        self.dimension = delegate.dimension
        self.cache: dict[str, np.ndarray] = {}

    def embed(self, texts: Sequence[str]) -> np.ndarray:
        missing = list(dict.fromkeys(text for text in texts if text not in self.cache))
        if missing:
            vectors = normalize_vectors(self.delegate.embed(missing))
            validate_dimension(vectors, self.dimension)
            self.cache.update(zip(missing, vectors))
        return np.asarray([self.cache[text] for text in texts], dtype=np.float32)


def normalize_vectors(vectors: np.ndarray) -> np.ndarray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    return vectors / np.where(norms == 0, 1.0, norms)


def validate_dimension(vectors: np.ndarray, dimension: int = DIMENSION) -> None:
    if vectors.ndim != 2 or vectors.shape[1] != dimension:
        raise ValueError(f"Expected embedding shape (*, {dimension}), got {vectors.shape}")


def create_embedder(config: dict[str, Any]) -> EmbeddingProvider:
    provider = config.get("provider", "sentence_transformers")
    if int(config.get("dimension", DIMENSION)) != DIMENSION:
        raise ValueError("This evaluation fixes BGE-M3 embedding dimension at 1024.")
    if provider == "sentence_transformers":
        delegate: EmbeddingProvider = SentenceTransformerProvider(config)
    elif provider == "http":
        delegate = HttpEmbeddingProvider(config)
    elif provider == "hash_test_only":
        delegate = HashEmbeddingProvider(DIMENSION)
    else:
        raise ValueError(f"Unsupported embedding provider: {provider}")
    return CachedEmbeddingProvider(delegate)


def embed_documents(
    provider: EmbeddingProvider,
    documents: Sequence[VectorDocument],
    batch_size: int,
) -> np.ndarray:
    unique_texts = list(dict.fromkeys(document.input_text for document in documents))
    cache: dict[str, np.ndarray] = {}
    for start in range(0, len(unique_texts), batch_size):
        batch = unique_texts[start : start + batch_size]
        vectors = normalize_vectors(provider.embed(batch))
        validate_dimension(vectors)
        cache.update(zip(batch, vectors))
    return np.asarray([cache[document.input_text] for document in documents], dtype=np.float32)


class VectorStore(Protocol):
    def replace_run(
        self,
        run_id: str,
        documents: Sequence[VectorDocument],
        vectors: np.ndarray,
    ) -> None: ...

    def search(
        self,
        run_id: str,
        variant: str,
        query_vector: np.ndarray,
        limit: int,
    ) -> list[SearchHit]: ...

    def close(self) -> None: ...


class InMemoryVectorStore:
    def __init__(self) -> None:
        self.documents: list[VectorDocument] = []
        self.vectors = np.empty((0, DIMENSION), dtype=np.float32)
        self.by_variant: dict[str, tuple[list[VectorDocument], np.ndarray]] = {}

    def replace_run(
        self, run_id: str, documents: Sequence[VectorDocument], vectors: np.ndarray
    ) -> None:
        del run_id
        self.documents = list(documents)
        self.vectors = normalize_vectors(np.asarray(vectors, dtype=np.float32))
        grouped_indexes: dict[str, list[int]] = defaultdict(list)
        for index, document in enumerate(self.documents):
            grouped_indexes[document.variant].append(index)
        self.by_variant = {
            variant: (
                [self.documents[index] for index in indexes],
                self.vectors[np.asarray(indexes, dtype=np.int64)],
            )
            for variant, indexes in grouped_indexes.items()
        }

    def search(
        self, run_id: str, variant: str, query_vector: np.ndarray, limit: int
    ) -> list[SearchHit]:
        del run_id
        variant_data = self.by_variant.get(variant)
        if not variant_data:
            return []
        documents, matrix = variant_data
        query_vector = normalize_vectors(np.asarray(query_vector, dtype=np.float32).reshape(1, -1))[0]
        scores = matrix @ query_vector
        order = np.argsort(-scores)[:limit]
        hits = []
        for position in order:
            document = documents[int(position)]
            hits.append(
                SearchHit(
                    entity_key=document.entity_key,
                    group_key=document.group_key,
                    entity_type=document.entity_type,
                    vector_role=document.vector_role,
                    similarity=float(scores[int(position)]),
                )
            )
        return hits

    def close(self) -> None:
        return


def _safe_identifier(value: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", value):
        raise ValueError(f"Unsafe SQL identifier: {value!r}")
    return value


class GaussVectorStore:
    """GaussDB/GaussVector floatvector(1024) adapter using cosine distance <+>."""

    def __init__(self, config: dict[str, Any]) -> None:
        try:
            import psycopg
        except ImportError as exc:  # pragma: no cover - dependency guidance
            raise RuntimeError("GaussVector backend requires psycopg[binary].") from exc
        dsn_env = str(config.get("dsn_env", "GAUSSVECTOR_DSN"))
        dsn = os.getenv(dsn_env)
        if not dsn:
            raise RuntimeError(f"Environment variable {dsn_env} is required.")
        self.psycopg = psycopg
        self.connection = psycopg.connect(dsn)
        self.schema = _safe_identifier(str(config.get("schema", "public")))
        self.table = _safe_identifier(str(config.get("table", "oag_vector_eval")))
        self.qualified = f'"{self.schema}"."{self.table}"'
        self.create_index = bool(config.get("create_index", True))
        raw_nlist = config.get("ivf_nlist", "auto")
        self.ivf_nlist = None if raw_nlist in {None, "auto"} else int(raw_nlist)
        self._ensure_table()

    def _ensure_table(self) -> None:
        ddl = f"""
        CREATE TABLE IF NOT EXISTS {self.qualified} (
          run_id VARCHAR(64) NOT NULL,
          variant VARCHAR(64) NOT NULL,
          entity_type VARCHAR(32) NOT NULL,
          entity_key VARCHAR(2048) NOT NULL,
          group_key VARCHAR(1024),
          vector_role VARCHAR(64) NOT NULL,
          segment_no INTEGER NOT NULL,
          input_text TEXT NOT NULL,
          embedding FLOATVECTOR({DIMENSION}) NOT NULL
        )
        """
        with self.connection.cursor() as cursor:
            cursor.execute(f'CREATE SCHEMA IF NOT EXISTS "{self.schema}"')
            cursor.execute(ddl)
        self.connection.commit()

    @staticmethod
    def _vector_literal(vector: np.ndarray) -> str:
        return "[" + ",".join(f"{float(value):.8g}" for value in vector) + "]"

    def replace_run(
        self, run_id: str, documents: Sequence[VectorDocument], vectors: np.ndarray
    ) -> None:
        validate_dimension(vectors)
        statement = f"""
        INSERT INTO {self.qualified}
          (run_id, variant, entity_type, entity_key, group_key,
           vector_role, segment_no, input_text, embedding)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, floatvector(%s))
        """
        rows = [
            (
                run_id,
                document.variant,
                document.entity_type,
                document.entity_key,
                document.group_key,
                document.vector_role,
                document.segment_no,
                document.input_text,
                self._vector_literal(vector),
            )
            for document, vector in zip(documents, vectors)
        ]
        with self.connection.cursor() as cursor:
            cursor.execute(f"DELETE FROM {self.qualified} WHERE run_id = %s", (run_id,))
            cursor.executemany(statement, rows)
        self.connection.commit()
        if self.create_index:
            self._ensure_index(len(documents))

    def _ensure_index(self, record_count: int) -> None:
        index_name = _safe_identifier(f"{self.table}_cosine_idx"[:60])
        ivf_nlist = self.ivf_nlist or max(1, math.ceil(4 * math.sqrt(record_count)))
        with self.connection.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM pg_indexes WHERE schemaname = %s AND indexname = %s",
                (self.schema, index_name),
            )
            exists = cursor.fetchone() is not None
            if not exists:
                cursor.execute(
                    f'CREATE INDEX "{index_name}" ON {self.qualified} '
                    f"USING gsivfflat (embedding cosine) WITH (ivf_nlist={ivf_nlist})"
                )
        self.connection.commit()

    def search(
        self, run_id: str, variant: str, query_vector: np.ndarray, limit: int
    ) -> list[SearchHit]:
        literal = self._vector_literal(query_vector)
        statement = f"""
        SELECT entity_key, group_key, entity_type, vector_role,
               embedding <+> floatvector(%s) AS distance
          FROM {self.qualified}
         WHERE run_id = %s AND variant = %s
         ORDER BY embedding <+> floatvector(%s)
         LIMIT %s
        """
        with self.connection.cursor() as cursor:
            cursor.execute(statement, (literal, run_id, variant, literal, limit))
            rows = cursor.fetchall()
        return [
            SearchHit(
                entity_key=row[0],
                group_key=row[1] or "",
                entity_type=row[2],
                vector_role=row[3],
                similarity=1.0 - float(row[4]),
            )
            for row in rows
        ]

    def close(self) -> None:
        self.connection.close()


def create_store(config: dict[str, Any]) -> VectorStore:
    backend = config.get("backend", "gaussvector")
    if backend == "gaussvector":
        return GaussVectorStore(config)
    if backend == "memory":
        return InMemoryVectorStore()
    raise ValueError(f"Unsupported vector_store.backend: {backend}")


def load_query_file(path: Path) -> list[QueryCase]:
    if not path.exists():
        return []
    cases: list[QueryCase] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            expected = tuple(part.strip() for part in row["expected_entity_keys"].split(";") if part.strip())
            units = tuple(part.strip() for part in row.get("semantic_units", "").split("|") if part.strip())
            if row.get("question", "").strip() and expected:
                cases.append(
                    QueryCase(
                        query_id=row.get("query_id") or f"file-{len(cases)+1}",
                        question=row["question"].strip(),
                        expected_entity_keys=expected,
                        semantic_units=units,
                        category=row.get("category", "file"),
                        language=row.get("language", "und"),
                        origin="file",
                    )
                )
    return cases


def auto_queries(
    entities: Sequence[EntityRecord], max_terms_per_entity: int, limit: int
) -> list[QueryCase]:
    term_to_keys: dict[str, set[str]] = defaultdict(set)
    term_original: dict[str, str] = {}
    term_categories: dict[str, set[str]] = defaultdict(set)
    for entity in entities:
        fields = entity.fields
        if entity.entity_type == "seed":
            candidates = [fields.get("name", ""), *split_synonyms(fields.get("synonyms", ""))]
        elif entity.entity_type == "metadata":
            candidates = split_synonyms(fields.get("synonyms", ""))
            if not candidates:
                candidates = [fields.get("name", ""), fields.get("value", "")]
        else:
            candidates = [fields.get("value", "")]
        for term in dedupe_parts(candidates)[:max_terms_per_entity]:
            normalized = normalize_term(term)
            if normalized:
                term_to_keys[normalized].add(entity.entity_key)
                term_original.setdefault(normalized, term)
                term_categories[normalized].add(entity.entity_type)

    cases: list[QueryCase] = []
    for normalized in sorted(term_to_keys):
        keys = tuple(sorted(term_to_keys[normalized]))
        cases.append(
            QueryCase(
                query_id=f"auto-{len(cases)+1:05d}",
                question=term_original[normalized],
                expected_entity_keys=keys,
                semantic_units=(term_original[normalized],),
                category="auto_" + "_".join(sorted(term_categories[normalized])),
                origin="auto",
            )
        )
        if limit and len(cases) >= limit:
            break
    return cases


def tokenize(question: str) -> list[str]:
    try:
        import jieba

        raw_tokens = list(jieba.cut(question, cut_all=False))
    except ImportError:
        expanded = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", question)
        raw_tokens = re.findall(r"[A-Za-z][A-Za-z0-9_-]*|[\u3400-\u9fff]{2,}|\d+", expanded)
    stopwords = {"的", "了", "和", "与", "或", "查询", "查找", "请", "用户"}
    return dedupe_parts(
        token for token in raw_tokens if normalize_term(token) not in stopwords and token.strip()
    )


def query_segments(case: QueryCase, strategy: str) -> list[str]:
    if strategy == "whole":
        return [case.question]
    if strategy == "token":
        return tokenize(case.question) or [case.question]
    if strategy == "semantic_phrase":
        return list(case.semantic_units) or [case.question]
    if strategy == "hybrid":
        return dedupe_parts([case.question, *(case.semantic_units or (case.question,))])
    raise ValueError(f"Unknown query strategy: {strategy}")


def rank_entities(
    store: VectorStore,
    provider: EmbeddingProvider,
    run_id: str,
    variant: str,
    segments: Sequence[str],
    document_limit: int,
    role_weights: dict[str, float],
) -> list[tuple[str, float]]:
    query_vectors = normalize_vectors(provider.embed(list(segments)))
    scores: dict[str, float] = {}
    for vector in query_vectors:
        for hit in store.search(run_id, variant, vector, document_limit):
            score = hit.similarity * float(role_weights.get(hit.vector_role, 1.0))
            scores[hit.entity_key] = max(scores.get(hit.entity_key, -math.inf), score)
    return sorted(scores.items(), key=lambda item: (-item[1], item[0]))


def dcg(relevances: Sequence[int]) -> float:
    return sum(value / math.log2(index + 2) for index, value in enumerate(relevances))


def percentile(values: Sequence[float], percentile_value: float) -> float:
    if not values:
        return 0.0
    return float(np.percentile(np.asarray(values), percentile_value))


def evaluate(
    store: VectorStore,
    provider: EmbeddingProvider,
    run_id: str,
    variant: str,
    query_strategy: str,
    cases: Sequence[QueryCase],
    top_ks: Sequence[int],
    document_limit: int,
    role_weights: dict[str, float],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    detail: list[dict[str, Any]] = []
    latencies: list[float] = []
    for case in cases:
        started = time.perf_counter()
        ranked = rank_entities(
            store,
            provider,
            run_id,
            variant,
            query_segments(case, query_strategy),
            document_limit,
            role_weights,
        )
        latency_ms = (time.perf_counter() - started) * 1000
        latencies.append(latency_ms)
        gold = set(case.expected_entity_keys)
        ranked_keys = [key for key, _ in ranked]
        gold_ranks = [ranked_keys.index(key) + 1 for key in gold if key in ranked_keys]
        best_rank = min(gold_ranks) if gold_ranks else None
        score_by_key = dict(ranked)
        best_positive = max((score_by_key.get(key, -1.0) for key in gold), default=-1.0)
        best_negative = max((score for key, score in ranked if key not in gold), default=-1.0)
        row: dict[str, Any] = {
            "variant": variant,
            "query_strategy": query_strategy,
            "query_id": case.query_id,
            "question": case.question,
            "category": case.category,
            "origin": case.origin,
            "gold_count": len(gold),
            "best_gold_rank": best_rank or "",
            "reciprocal_rank": 1.0 / best_rank if best_rank else 0.0,
            "best_positive_similarity": best_positive,
            "best_negative_similarity": best_negative,
            "similarity_margin": best_positive - best_negative,
            "latency_ms": latency_ms,
            "top1_entity_key": ranked[0][0] if ranked else "",
            "top1_similarity": ranked[0][1] if ranked else "",
            "top10": json.dumps(ranked[:10], ensure_ascii=False),
        }
        for k in top_ks:
            selected = set(ranked_keys[:k])
            relevant_count = len(selected & gold)
            row[f"hit@{k}"] = float(relevant_count > 0)
            row[f"recall@{k}"] = relevant_count / len(gold) if gold else 0.0
            row[f"precision@{k}"] = relevant_count / k
            relevances = [1 if key in gold else 0 for key in ranked_keys[:k]]
            ideal = [1] * min(k, len(gold))
            row[f"ndcg@{k}"] = dcg(relevances) / dcg(ideal) if ideal else 0.0
        detail.append(row)

    summary: dict[str, Any] = {
        "variant": variant,
        "query_strategy": query_strategy,
        "query_count": len(detail),
        "mrr": statistics.fmean(float(row["reciprocal_rank"]) for row in detail) if detail else 0.0,
        "mean_best_positive_similarity": statistics.fmean(
            float(row["best_positive_similarity"]) for row in detail
        ) if detail else 0.0,
        "mean_similarity_margin": statistics.fmean(float(row["similarity_margin"]) for row in detail)
        if detail
        else 0.0,
        "latency_p50_ms": percentile(latencies, 50),
        "latency_p95_ms": percentile(latencies, 95),
    }
    for k in top_ks:
        for metric in ("hit", "recall", "precision", "ndcg"):
            name = f"{metric}@{k}"
            summary[name] = statistics.fmean(float(row[name]) for row in detail) if detail else 0.0
    return summary, detail


def write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    columns = list(rows[0])
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def document_manifest(documents: Sequence[VectorDocument]) -> list[dict[str, Any]]:
    grouped: dict[str, list[VectorDocument]] = defaultdict(list)
    for document in documents:
        grouped[document.variant].append(document)
    rows = []
    for variant, items in sorted(grouped.items()):
        lengths = [len(item.input_text) for item in items]
        rows.append(
            {
                "variant": variant,
                "vector_rows": len(items),
                "unique_entities": len({item.entity_key for item in items}),
                "mean_input_chars": statistics.fmean(lengths) if lengths else 0,
                "p95_input_chars": percentile(lengths, 95),
                "max_input_chars": max(lengths, default=0),
            }
        )
    return rows


def write_markdown_report(
    path: Path,
    run_id: str,
    summaries: Sequence[dict[str, Any]],
    manifest: Sequence[dict[str, Any]],
    quality: DataQuality,
    top_ks: Sequence[int],
    embedding_config: dict[str, Any],
    vector_store_config: dict[str, Any],
) -> None:
    primary_k = 1 if 1 in top_ks else min(top_ks)
    ranked = sorted(
        summaries,
        key=lambda row: (-float(row[f"hit@{primary_k}"]), -float(row["mrr"]), float(row["latency_p95_ms"])),
    )
    lines = [
        "# OAG向量化结构检索评测结果",
        "",
        f"- run_id: `{run_id}`",
        f"- embedding provider: `{embedding_config.get('provider')}`",
        f"- embedding model: `{embedding_config.get('model_name_or_path', 'BAAI/bge-m3')}`，1024维，L2归一化",
        f"- vector backend: `{vector_store_config.get('backend')}`",
        "- distance: COSINE（GaussVector模式使用`<+>`）",
        "",
        "## 指标对比",
        "",
        f"|排名|向量方案|查询策略|Hit@{primary_k}|MRR|Recall@10|平均正样本相似度|平均相似度间隔|P95延迟(ms)|",
        "|---:|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for index, row in enumerate(ranked, start=1):
        lines.append(
            f"|{index}|{row['variant']}|{row['query_strategy']}|"
            f"{float(row[f'hit@{primary_k}']):.4f}|{float(row['mrr']):.4f}|"
            f"{float(row.get('recall@10', 0)):.4f}|"
            f"{float(row['mean_best_positive_similarity']):.4f}|"
            f"{float(row['mean_similarity_margin']):.4f}|"
            f"{float(row['latency_p95_ms']):.2f}|"
        )
    lines.extend(["", "## 索引成本", "", "|向量方案|向量行数|实体数|平均字符数|P95字符数|最大字符数|", "|---|---:|---:|---:|---:|---:|"])
    for row in manifest:
        lines.append(
            f"|{row['variant']}|{row['vector_rows']}|{row['unique_entities']}|"
            f"{float(row['mean_input_chars']):.1f}|{float(row['p95_input_chars']):.1f}|{row['max_input_chars']}|"
        )
    lines.extend(
        [
            "",
            "## 数据质量",
            "",
            f"```json\n{json.dumps(asdict(quality), ensure_ascii=False, indent=2)}\n```",
            "",
            "> 自动生成问题只适合管线基线；最终方案必须以专家标注的真实问题集为准。",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def deep_update(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    output = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(output.get(key), dict):
            output[key] = deep_update(output[key], value)
        else:
            output[key] = value
    return output


def default_config() -> dict[str, Any]:
    return {
        "input": {
            "workbook": "tests/OAG_本体索引结构.xlsx",
            "query_file": "tests/oag_vector_eval_queries.csv",
            "auto_generate_queries": True,
            "auto_query_limit": 2000,
            "max_auto_terms_per_entity": 3,
        },
        "embedding": {
            "provider": "sentence_transformers",
            "model_name_or_path": "BAAI/bge-m3",
            "dimension": DIMENSION,
            "batch_size": 32,
            "normalize_embeddings": True,
            "device": None,
        },
        "vector_store": {
            "backend": "gaussvector",
            "dsn_env": "GAUSSVECTOR_DSN",
            "schema": "public",
            "table": "oag_vector_eval",
            "create_index": True,
            "ivf_nlist": "auto",
        },
        "evaluation": {
            "variants": DEFAULT_VARIANTS,
            "query_strategies": DEFAULT_QUERY_STRATEGIES,
            "top_ks": [1, 3, 5, 10],
            "document_limit": 100,
            "role_weights": {"description": 0.95, "context": 0.95},
        },
        "output": {"directory": "tests/oag_vector_eval_output"},
    }


def run(config: dict[str, Any]) -> dict[str, Any]:
    input_config = config["input"]
    entities, quality = load_workbook(Path(input_config["workbook"]))
    variants = list(config["evaluation"]["variants"])
    documents = [document for variant in variants for document in build_documents(entities, variant)]
    provider = create_embedder(config["embedding"])
    vectors = embed_documents(provider, documents, int(config["embedding"].get("batch_size", 32)))
    store = create_store(config["vector_store"])
    run_id = str(config.get("run_id") or uuid.uuid4().hex[:16])
    output_dir = Path(config["output"]["directory"])
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        store.replace_run(run_id, documents, vectors)
        cases = load_query_file(Path(input_config["query_file"]))
        if input_config.get("auto_generate_queries", True):
            cases.extend(
                auto_queries(
                    entities,
                    int(input_config.get("max_auto_terms_per_entity", 3)),
                    int(input_config.get("auto_query_limit", 2000)),
                )
            )
        unique_cases = {case.query_id: case for case in cases}
        cases = list(unique_cases.values())
        if not cases:
            raise ValueError("No labeled queries were loaded or generated.")

        summaries: list[dict[str, Any]] = []
        details: list[dict[str, Any]] = []
        for variant in variants:
            for query_strategy in config["evaluation"]["query_strategies"]:
                summary, rows = evaluate(
                    store,
                    provider,
                    run_id,
                    variant,
                    query_strategy,
                    cases,
                    config["evaluation"]["top_ks"],
                    int(config["evaluation"].get("document_limit", 100)),
                    config["evaluation"].get("role_weights", {}),
                )
                summaries.append(summary)
                details.extend(rows)
    finally:
        store.close()

    manifest = document_manifest(documents)
    write_csv(output_dir / "summary.csv", summaries)
    write_csv(output_dir / "query_results.csv", details)
    write_csv(output_dir / "index_manifest.csv", manifest)
    (output_dir / "data_quality.json").write_text(
        json.dumps(asdict(quality), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_markdown_report(
        output_dir / "summary.md",
        run_id,
        summaries,
        manifest,
        quality,
        config["evaluation"]["top_ks"],
        config["embedding"],
        config["vector_store"],
    )
    result = {"run_id": run_id, "output_directory": str(output_dir), "query_count": len(cases)}
    (output_dir / "run.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, help="JSON configuration file")
    parser.add_argument("--workbook", type=Path, help="Override workbook path")
    parser.add_argument("--query-file", type=Path, help="Override labeled query CSV")
    parser.add_argument("--output-dir", type=Path, help="Override output directory")
    parser.add_argument("--backend", choices=["gaussvector", "memory"], help="Override vector backend")
    parser.add_argument(
        "--embedding-provider",
        choices=["sentence_transformers", "http", "hash_test_only"],
        help="Override embedding provider",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = default_config()
    if args.config:
        config = deep_update(config, json.loads(args.config.read_text(encoding="utf-8")))
    if args.workbook:
        config["input"]["workbook"] = str(args.workbook)
    if args.query_file:
        config["input"]["query_file"] = str(args.query_file)
    if args.output_dir:
        config["output"]["directory"] = str(args.output_dir)
    if args.backend:
        config["vector_store"]["backend"] = args.backend
    if args.embedding_provider:
        config["embedding"]["provider"] = args.embedding_provider
    print(json.dumps(run(config), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

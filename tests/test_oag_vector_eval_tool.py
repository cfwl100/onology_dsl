from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

import oag_vector_eval as tool


def make_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    seed = workbook.active
    seed.title = "种子节点"
    seed.append(
        [
            "type",
            "id",
            "parent_id",
            "name",
            "display_zh",
            "display_en",
            "display_lang_1",
            "display_lang_2",
            "description_zh",
            "description_en",
            "description_lang_1",
            "description_lang_2",
            "synonyms",
        ]
    )
    seed.append([0, "object-account", None, "Account", None, None, None, None, "账户", "Account", None, None, "账户\nAccount"])
    seed.append([1, "property-status", "object-account", "status", None, None, None, None, "账户状态", "Account status", None, None, "状态\nStatus"])

    metadata = workbook.create_sheet("元数据元素")
    metadata.append(
        [
            "type",
            "propertyId",
            "objectTypeId",
            "value",
            "name",
            "display_zh",
            "display_en",
            "display_lang_1",
            "display_lang_2",
            "description_zh",
            "description_en",
            "description_lang_1",
            "description_lang_2",
            "synonyms",
        ]
    )
    metadata.append([2, "enum-AccountStatus", "object-account", "1", "1", None, None, None, None, "有效账户", "Active account", None, None, "有效\nActive"])

    instance = workbook.create_sheet("实例元素")
    instance.append(["type", "propertyid", "objectTypeId", "value"])
    instance.append([3, "property-status", "object-account", None])
    workbook.save(path)


def test_loader_and_document_variants(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fixture.xlsx"
    make_workbook(workbook_path)
    entities, quality = tool.load_workbook(workbook_path)

    assert len(entities) == 3
    assert quality.loaded_entities == {"seed": 2, "metadata": 1}
    assert quality.skipped_empty_values == {"instance.value": 1}
    assert len(tool.build_documents(entities, "all_fields_single")) == 3
    assert len(tool.build_documents(entities, "core_short_single")) == 3
    assert len(tool.build_documents(entities, "term_description_multi")) == 6
    assert len(tool.build_documents(entities, "context_shadow_multi")) > 6
    assert len(tool.build_documents(entities, "atomic_field_rows")) > 6


def test_query_segmentation() -> None:
    case = tool.QueryCase(
        query_id="q1",
        question="查询FORMAL用户的Mobile Number",
        expected_entity_keys=("seed|subscriber",),
        semantic_units=("FORMAL用户", "Mobile Number"),
    )
    assert tool.query_segments(case, "whole") == [case.question]
    assert tool.query_segments(case, "semantic_phrase") == ["FORMAL用户", "Mobile Number"]
    assert tool.query_segments(case, "hybrid")[0] == case.question


def test_end_to_end_memory_pipeline(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fixture.xlsx"
    query_path = tmp_path / "queries.csv"
    output_path = tmp_path / "output"
    make_workbook(workbook_path)
    query_path.write_text(
        "query_id,question,expected_entity_keys,semantic_units,category,language\n"
        "q1,账户状态,seed|property-status,账户状态,seed_property,zh\n"
        "q2,有效账户,metadata|enum-AccountStatus|object-account|1,有效账户,enum_value,zh\n",
        encoding="utf-8",
    )
    config = tool.default_config()
    config["input"].update(
        {
            "workbook": str(workbook_path),
            "query_file": str(query_path),
            "auto_generate_queries": False,
        }
    )
    config["embedding"]["provider"] = "hash_test_only"
    config["vector_store"]["backend"] = "memory"
    config["evaluation"]["variants"] = ["all_fields_single", "term_description_multi"]
    config["evaluation"]["query_strategies"] = ["whole", "semantic_phrase"]
    config["output"]["directory"] = str(output_path)

    result = tool.run(config)

    assert result["query_count"] == 2
    assert (output_path / "summary.csv").exists()
    assert (output_path / "query_results.csv").exists()
    assert (output_path / "summary.md").exists()
    assert json.loads((output_path / "data_quality.json").read_text(encoding="utf-8"))[
        "skipped_empty_values"
    ] == {"instance.value": 1}
